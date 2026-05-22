#!/usr/bin/env python3
import subprocess
import json
import re
from datetime import datetime
from openpyxl import load_workbook

SHEET_ID = "1ECqrsi3HjcQxEf_FVVQHDJppwUga_-pHRjRpRLNyOrQ"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
XLSX_FILE = "/tmp/mcl_data.xlsx"
OUTPUT_FILE = "data.json"

def download_sheets():
    print(f"Downloading Sheets...")
    result = subprocess.run(["curl", "-L", "-o", XLSX_FILE, EXPORT_URL], capture_output=True)
    return result.returncode == 0

def extract_value(cell_val):
    if not cell_val or not isinstance(cell_val, str): return cell_val
    match = re.search(r'"""COMPUTED_VALUE"""\),(.+)\)', cell_val)
    if match:
        try: return json.loads(match.group(1))
        except: return cell_val
    return cell_val

def excel_date_to_str(serial):
    if not serial or not isinstance(serial, (int, float)): return None
    try:
        from datetime import datetime, timedelta
        date_obj = datetime(1899, 12, 30) + timedelta(days=serial)
        return date_obj.strftime("%Y/%m/%d")
    except: return None

def parse_sheets():
    wb = load_workbook(XLSX_FILE)
    data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "定義": continue
        ws = wb[sheet_name]
        rows = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for i, cell in enumerate(row):
                if i >= len(headers): break
                val = cell.value
                if isinstance(val, str) and "COMPUTED_VALUE" in val:
                    val = extract_value(val)
                if isinstance(val, (int, float)) and 40000 < val < 50000:
                    val = excel_date_to_str(val)
                row_data[headers[i]] = val
            if any(row_data.values()): rows.append(row_data)
        if rows: data[sheet_name] = rows
    return data

def main():
    print(f"[{datetime.now().isoformat()}] Updating MCL dashboard data...")
    if not download_sheets():
        print("Failed to download sheets")
        return False
    print("Parsing XLSX...")
    data = parse_sheets()
    output_data = {}
    for sheet_name, rows in data.items():
        for row in rows:
            date_str = None
            if sheet_name == "プラン比率" and "申込日" in row:
                date_str = row["申込日"]
            elif "発生日" in row:
                date_str = row["発生日"]
            elif "来店日" in row:
                date_str = row["来店日"]
            month_key = date_str[:7] if date_str and isinstance(date_str, str) and len(date_str) >= 7 else "2026-05"
            if month_key not in output_data: output_data[month_key] = {}
            if sheet_name not in output_data[month_key]:
                output_data[month_key][sheet_name] = []
            output_data[month_key][sheet_name].append(row)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump({"ALL_DATA": output_data}, f, ensure_ascii=False)
    print(f"✓ Saved to {OUTPUT_FILE} ({len(output_data)} months)")
    return True

if __name__ == "__main__":
    exit(0 if main() else 1)
